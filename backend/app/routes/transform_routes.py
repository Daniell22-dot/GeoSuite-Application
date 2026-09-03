"""
Coordinate transformation routes.
Cassini-Soldner ↔ UTM conversions for Kenya survey plans.
"""
from fastapi import APIRouter, UploadFile, File, HTTPException
from pydantic import BaseModel
from typing import List, Optional, Dict
import os
import tempfile
import shutil

from app.services.transform_service import (
    CoordinateTransformer,
    PolynomialCoefficients,
    KENYA_CASSINI_ZONES,
)

router = APIRouter()
transformer = CoordinateTransformer()


# ── Request models ──────────────────────────────────────────────────

class SingleTransformRequest(BaseModel):
    easting: float
    northing: float
    direction: str = "cassini_to_utm"  # or "utm_to_cassini"
    cassini_zone: str = "zone_3"
    utm_zone: int = 37
    method: str = "geodetic"  # geodetic, polynomial, helmert
    polynomial_coeffs: Optional[Dict[str, float]] = None


class BulkTransformRequest(BaseModel):
    coordinates: List[Dict[str, float]]  # [{"easting": ..., "northing": ...}, ...]
    direction: str = "cassini_to_utm"
    cassini_zone: str = "zone_3"
    utm_zone: int = 37
    method: str = "geodetic"
    polynomial_coeffs: Optional[Dict[str, float]] = None


class PolynomialUploadRequest(BaseModel):
    easting: float
    northing: float
    a0: float = 0.0
    a1: float = 1.0
    a2: float = 0.0
    b0: float = 0.0
    b1: float = 0.0
    b2: float = 1.0


# ── Endpoints ───────────────────────────────────────────────────────

@router.post("/single")
async def transform_single(request: SingleTransformRequest):
    """
    Transform a single coordinate pair.

    **Methods:**
    - `geodetic`: Proper geodetic chain via pyproj (most accurate, no control points needed)
    - `polynomial`: Affine/polynomial transformation (matches Excel formulae from survey offices)
    - `helmert`: 7-parameter Helmert transformation (requires known control points)
    """
    if request.method == "polynomial":
        if not request.polynomial_coeffs:
            raise HTTPException(
                status_code=400,
                detail="polynomial_coeffs required for polynomial method"
            )
        coeffs = PolynomialCoefficients(**request.polynomial_coeffs)
        result = transformer.cassini_to_utm_polynomial(
            request.easting, request.northing, coeffs
        )
    elif request.method == "helmert":
        from app.services.transform_service import KENYA_DATUM_SHIFTS
        result = transformer.cassini_to_utm_helmert(
            request.easting, request.northing, KENYA_DATUM_SHIFTS["arc_1960"]
        )
    elif request.direction == "utm_to_cassini":
        result = transformer.utm_to_cassini_geodetic(
            request.easting, request.northing,
            request.cassini_zone, request.utm_zone,
        )
    else:
        result = transformer.cassini_to_utm_geodetic(
            request.easting, request.northing,
            request.cassini_zone, request.utm_zone,
        )

    if not result:
        raise HTTPException(status_code=422, detail="Transformation failed")

    return {
        "input": {"easting": result.input_easting, "northing": result.input_northing},
        "output": {"easting": result.output_easting, "northing": result.output_northing},
        "geographic": {"latitude": result.latitude, "longitude": result.longitude},
        "method": result.method,
        "output_epsg": result.output_epsg,
        "accuracy_m": result.accuracy_m,
    }


@router.post("/bulk")
async def transform_bulk(request: BulkTransformRequest):
    """
    Transform multiple coordinate pairs at once.
    Pass up to 10,000 coordinates in a single request.
    """
    if len(request.coordinates) > 10000:
        raise HTTPException(status_code=400, detail="Maximum 10,000 coordinates per request")

    results = []
    errors = []

    for i, coord in enumerate(request.coordinates):
        try:
            E = float(coord.get("easting", 0))
            N = float(coord.get("northing", 0))

            if request.method == "polynomial":
                if request.polynomial_coeffs:
                    coeffs = PolynomialCoefficients(**request.polynomial_coeffs)
                else:
                    errors.append({"index": i, "error": "No polynomial coefficients"})
                    continue
                result = transformer.cassini_to_utm_polynomial(E, N, coeffs)
            elif request.direction == "utm_to_cassini":
                result = transformer.utm_to_cassini_geodetic(
                    E, N, request.cassini_zone, request.utm_zone
                )
            else:
                result = transformer.cassini_to_utm_geodetic(
                    E, N, request.cassini_zone, request.utm_zone
                )

            if result:
                results.append({
                    "index": i,
                    "input": {"easting": E, "northing": N},
                    "output": {"easting": result.output_easting, "northing": result.output_northing},
                    "latitude": result.latitude,
                    "longitude": result.longitude,
                })
            else:
                errors.append({"index": i, "error": "Transformation failed"})

        except Exception as e:
            errors.append({"index": i, "error": str(e)})

    return {
        "total": len(request.coordinates),
        "successful": len(results),
        "failed": len(errors),
        "results": results,
        "errors": errors,
    }


@router.post("/upload-excel")
async def transform_from_excel(
    file: UploadFile = File(...),
    method: str = "geodetic",
    cassini_zone: str = "zone_3",
    utm_zone: int = 37,
):
    """
    Upload an Excel file with Cassini coordinates and get UTM results.

    **Supported column headers** (auto-detected):
    - Easting: `easting`, `east`, `x`, `e_cassini`, `c_east`
    - Northing: `northing`, `north`, `y`, `n_cassini`, `c_north`

    **Polynomial coefficients:** If using the polynomial method, either:
    - Put coefficients in a second sheet with labels (a0, a1, a2, b0, b1, b2)
    - Or pass them in the request body

    Returns a downloadable Excel file with the transformed coordinates.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in ('.xlsx', '.xls'):
        raise HTTPException(status_code=400, detail="Only .xlsx files supported")

    # Save uploaded file
    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, file.filename)

    try:
        content = await file.read()
        with open(input_path, "wb") as f:
            f.write(content)

        # Transform
        result = transformer.transform_excel(
            input_path,
            method=method,
            zone=cassini_zone,
            utm_zone=utm_zone,
        )

        if "error" in result:
            raise HTTPException(status_code=422, detail=result["error"])

        # Generate output Excel
        output_path = os.path.join(temp_dir, "transformed_coordinates.xlsx")
        _write_results_to_excel(result, output_path)

        from fastapi.responses import FileResponse
        return FileResponse(
            output_path,
            filename="transformed_coordinates.xlsx",
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"X-Transform-Stats": f"{result['successful']}/{result['total_rows']} successful"},
        )

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@router.post("/upload-excel-preview")
async def transform_excel_preview(
    file: UploadFile = File(...),
    method: str = "geodetic",
    cassini_zone: str = "zone_3",
    utm_zone: int = 37,
):
    """
    Preview transformation results without downloading.
    Returns first 100 rows transformed.
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    temp_dir = tempfile.mkdtemp()
    input_path = os.path.join(temp_dir, file.filename)

    try:
        content = await file.read()
        with open(input_path, "wb") as f:
            f.write(content)

        result = transformer.transform_excel(
            input_path,
            method=method,
            zone=cassini_zone,
            utm_zone=utm_zone,
        )

        # Limit preview to 100 rows
        if "results" in result:
            result["results"] = result["results"][:100]
            result["preview"] = True

        return result

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)


@router.get("/zones")
async def get_zones():
    """List all supported Cassini zones and transformation methods."""
    return transformer.get_available_zones()


@router.get("/detect-zone")
async def detect_zone(longitude: float):
    """Auto-detect the best Cassini and UTM zones from a longitude."""
    return {
        "longitude": longitude,
        "cassini_zone": transformer.detect_cassini_zone(longitude),
        "utm_zone": transformer.detect_utm_zone(longitude),
        "cassini_zones": KENYA_CASSINI_ZONES,
    }


# ── Helper ──────────────────────────────────────────────────────────

def _write_results_to_excel(result: Dict, output_path: str):
    """Write transformation results to an Excel file."""
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Transformed Coordinates"

    # Headers
    headers = [
        "Input Easting", "Input Northing",
        "Output Easting (UTM)", "Output Northing (UTM)",
        "Latitude", "Longitude",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for row_data in result.get("results", []):
        row = row_data.get("row", row_data.get("index", 0)) + 1
        input_data = row_data.get("input", {})
        output_data = row_data.get("output", {})

        ws.cell(row=row, column=1, value=input_data.get("easting", 0))
        ws.cell(row=row, column=2, value=input_data.get("northing", 0))
        ws.cell(row=row, column=3, value=output_data.get("easting", 0))
        ws.cell(row=row, column=4, value=output_data.get("northing", 0))
        ws.cell(row=row, column=5, value=row_data.get("latitude", ""))
        ws.cell(row=row, column=6, value=row_data.get("longitude", ""))

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Method")
    ws2.cell(row=1, column=2, value=result.get("method", "unknown"))
    ws2.cell(row=2, column=1, value="Total Rows")
    ws2.cell(row=2, column=2, value=result.get("total_rows", 0))
    ws2.cell(row=3, column=1, value="Successful")
    ws2.cell(row=3, column=2, value=result.get("successful", 0))
    ws2.cell(row=4, column=1, value="Failed")
    ws2.cell(row=4, column=2, value=result.get("failed", 0))
    ws2.cell(row=5, column=1, value="Cassini Zone")
    ws2.cell(row=5, column=2, value=result.get("zone", ""))
    ws2.cell(row=6, column=1, value="UTM Zone")
    ws2.cell(row=6, column=2, value=result.get("utm_zone", ""))

    wb.save(output_path)
