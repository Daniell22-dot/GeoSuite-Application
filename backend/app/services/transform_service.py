"""
Coordinate transformation service.
Handles Kenya Cassini-Soldner ↔ UTM conversions via:
  1. Proper geodetic chain (pyproj + custom CRS definitions)
  2. Polynomial/affine transformations (matching Excel formulae used by surveyors)
  3. Bulk transformation from Excel/CSV files

Kenya uses Cassini-Soldner with Clarke 1880 ellipsoid on legacy survey plans.
Modern mapping uses UTM Zone 37S/36S on WGS84.
"""
import os
import math
import json
from typing import Dict, List, Optional, Tuple, Union
from dataclasses import dataclass, field

import numpy as np

try:
    from pyproj import Transformer, CRS
    from pyproj.aoi import AreaOfInterest
    from pyproj.database import query_utm_crs_info
    PYPROJ_AVAILABLE = True
except ImportError:
    PYPROJ_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ── Kenya-specific CRS definitions ──────────────────────────────────

# Kenya Cassini-Soldner zones (Clarke 1880 RGS ellipsoid)
KENYA_CASSINI_ZONES = {
    "zone_1": {
        "name": "Cassini Zone I",
        "central_meridian": 34.0,   # 34°E
        "false_easting": 700000.0,
        "false_northing": 5000000.0,
        "latitude_of_origin": 0.0,  # Equator
        "description": "Western Kenya (Lake Victoria region)",
    },
    "zone_2": {
        "name": "Cassini Zone II",
        "central_meridian": 36.0,   # 36°E
        "false_easting": 700000.0,
        "false_northing": 5000000.0,
        "latitude_of_origin": 0.0,
        "description": "Central-West Kenya",
    },
    "zone_3": {
        "name": "Cassini Zone III",
        "central_meridian": 38.0,   # 38°E
        "false_easting": 700000.0,
        "false_northing": 5000000.0,
        "latitude_of_origin": 0.0,
        "description": "Central Kenya (Nairobi region)",
    },
    "zone_4": {
        "name": "Cassini Zone IV",
        "central_meridian": 40.0,   # 40°E
        "false_easting": 700000.0,
        "false_northing": 5000000.0,
        "latitude_of_origin": 0.0,
        "description": "Eastern Kenya (Coast region)",
    },
    "nairobi": {
        "name": "Nairobi Local Cassini",
        "central_meridian": 36.8233,  # 36°49'23.378"E
        "false_easting": 700000.0,
        "false_northing": 5000000.0,
        "latitude_of_origin": -1.3737,  # 1°22'25.187"S
        "description": "Nairobi local survey grid",
    },
}

# Clarke 1880 RGS ellipsoid (used in Kenya)
CLARKE_1880_RGS = {
    "a": 6378249.145,      # semi-major axis (metres)
    "rf": 293.4663,         # inverse flattening
}

# WGS84 ellipsoid
WGS84 = {
    "a": 6378137.0,
    "rf": 298.257223563,
}

# Common datum shifts used in Kenya (Clarke 1880 → WGS84)
# These are approximate 3-parameter (Bursa-Wolf) transforms
KENYA_DATUM_SHIFTS = {
    "arc_1960": {
        "name": "Arc 1960 (Kenya/Tanzania)",
        "dx": -156.0,
        "dy": -8.0,
        "dz": -294.0,
    },
    "asia_north": {
        "name": "Asia North Datum",
        "dx": -163.0,
        "dy": -6.0,
        "dz": -289.0,
    },
}


@dataclass
class TransformResult:
    """Result of a coordinate transformation."""
    input_easting: float
    input_northing: float
    input_zone: str
    output_easting: float
    output_northing: float
    output_epsg: int
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    method: str = "geodetic"
    accuracy_m: Optional[float] = None


@dataclass
class PolynomialCoefficients:
    """Polynomial transformation coefficients (matching Excel formulae)."""
    # E_utm = a0 + a1*E_cass + a2*N_cass + a3*E² + a4*E*N + a5*N²
    a0: float = 0.0
    a1: float = 1.0
    a2: float = 0.0
    a3: float = 0.0
    a4: float = 0.0
    a5: float = 0.0
    # N_utm = b0 + b1*E_cass + b2*N_cass + b3*E² + b4*E*N + b5*N²
    b0: float = 0.0
    b1: float = 0.0
    b2: float = 1.0
    b3: float = 0.0
    b4: float = 0.0
    b5: float = 0.0


class CoordinateTransformer:
    """
    Transforms coordinates between Kenya Cassini-Soldner and UTM.

    Three approaches:
    1. Geodetic chain via pyproj (most accurate, no control points needed)
    2. Polynomial transformation (matches Excel formulae from survey offices)
    3. Helmert transformation (7-parameter, requires known control points)
    """

    def __init__(self):
        self.pyproj_available = PYPROJ_AVAILABLE
        self._transformers = {}

    # ══════════════════════════════════════════════════════════════════
    #  APPROACH 1: Proper geodetic chain via pyproj
    # ══════════════════════════════════════════════════════════════════

    def _get_cassini_crs(self, zone: str) -> Optional[CRS]:
        """Build a pyproj CRS for Kenya Cassini-Soldner."""
        if not self.pyproj_available:
            return None

        zone_params = KENYA_CASSINI_ZONES.get(zone)
        if not zone_params:
            return None

        cache_key = f"cassini_{zone}"
        if cache_key in self._transformers:
            return self._transformers[cache_key]

        # Build custom CRS using PROJ string
        # Cassini-Soldner on Clarke 1880 RGS
        proj4 = (
            f"+proj=cass +lat_0={zone_params['latitude_of_origin']} "
            f"+lon_0={zone_params['central_meridian']} "
            f"+k=1 "
            f"+x_0={zone_params['false_easting']} "
            f"+y_0={zone_params['false_northing']} "
            f"+ellps=clrk80 "
            f"+towgs84=-163,-6,-289,0,0,0,0 "  # Arc 1960 approximate shift
            f"+units=m +no_defs"
        )

        try:
            crs = CRS.from_proj4(proj4)
            self._transformers[cache_key] = crs
            return crs
        except Exception:
            return None

    def _get_utm_crs(self, zone: int = 37, south: bool = True) -> Optional[CRS]:
        """Get UTM CRS for Kenya."""
        if not self.pyproj_available:
            return None

        epsg = 32700 + zone if south else 32600 + zone
        cache_key = f"utm_{epsg}"

        if cache_key in self._transformers:
            return self._transformers[cache_key]

        try:
            crs = CRS.from_epsg(epsg)
            self._transformers[cache_key] = crs
            return crs
        except Exception:
            return None

    def cassini_to_utm_geodetic(
        self,
        easting: float,
        northing: float,
        zone: str = "zone_3",
        utm_zone: int = 37,
    ) -> Optional[TransformResult]:
        """
        Convert Cassini-Soldner → UTM via proper geodetic chain:
          Cassini (Clarke 1880) → Geographic → WGS84 → UTM

        Args:
            easting: Cassini Easting in metres
            northing: Cassini Northing in metres
            zone: Kenya Cassini zone key
            utm_zone: Target UTM zone (36 or 37 for Kenya)

        Returns:
            TransformResult with UTM coordinates
        """
        cassini_crs = self._get_cassini_crs(zone)
        utm_crs = self._get_utm_crs(utm_zone)

        if not cassini_crs or not utm_crs:
            return None

        try:
            transformer = Transformer.from_crs(cassini_crs, utm_crs, always_xy=True)
            utm_e, utm_n = transformer.transform(easting, northing)

            # Also get geographic coordinates
            geo_transformer = Transformer.from_crs(
                cassini_crs, "EPSG:4326", always_xy=True
            )
            lon, lat = geo_transformer.transform(easting, northing)

            return TransformResult(
                input_easting=easting,
                input_northing=northing,
                input_zone=zone,
                output_easting=round(utm_e, 3),
                output_northing=round(utm_n, 3),
                output_epsg=32700 + utm_zone,
                latitude=round(lat, 8),
                longitude=round(lon, 8),
                method="geodetic_chain",
                accuracy_m=1.0,  # ~1m accuracy with Arc 1960 shift
            )
        except Exception as e:
            return None

    def utm_to_cassini_geodetic(
        self,
        easting: float,
        northing: float,
        zone: str = "zone_3",
        utm_zone: int = 37,
    ) -> Optional[TransformResult]:
        """Convert UTM → Cassini-Soldner (reverse of above)."""
        cassini_crs = self._get_cassini_crs(zone)
        utm_crs = self._get_utm_crs(utm_zone)

        if not cassini_crs or not utm_crs:
            return None

        try:
            transformer = Transformer.from_crs(utm_crs, cassini_crs, always_xy=True)
            cass_e, cass_n = transformer.transform(easting, northing)

            geo_transformer = Transformer.from_crs(
                utm_crs, "EPSG:4326", always_xy=True
            )
            lon, lat = geo_transformer.transform(easting, northing)

            return TransformResult(
                input_easting=easting,
                input_northing=northing,
                input_zone=f"utm_{utm_zone}",
                output_easting=round(cass_e, 3),
                output_northing=round(cass_n, 3),
                output_epsg=0,  # Custom CRS
                latitude=round(lat, 8),
                longitude=round(lon, 8),
                method="geodetic_chain_reverse",
                accuracy_m=1.0,
            )
        except Exception:
            return None

    # ══════════════════════════════════════════════════════════════════
    #  APPROACH 2: Polynomial / Affine (matching Excel formulae)
    # ══════════════════════════════════════════════════════════════════

    def cassini_to_utm_polynomial(
        self,
        easting: float,
        northing: float,
        coeffs: PolynomialCoefficients,
    ) -> TransformResult:
        """
        Convert Cassini → UTM using polynomial transformation.
        This matches the Excel formulae used by Kenya surveyors.

        Formula (affine):
            E_utm = a0 + a1*E_cass + a2*N_cass
            N_utm = b0 + b1*E_cass + b2*N_cass

        Formula (2nd order polynomial):
            E_utm = a0 + a1*E + a2*N + a3*E² + a4*E*N + a5*N²
            N_utm = b0 + b1*E + b2*N + b3*E² + b4*E*N + b5*N²
        """
        E = easting
        N = northing

        # Check if any 2nd-order terms are non-zero
        has_quadratic = any([
            coeffs.a3, coeffs.a4, coeffs.a5,
            coeffs.b3, coeffs.b4, coeffs.b5,
        ])

        if has_quadratic:
            # Full 2nd-order polynomial
            utm_e = (coeffs.a0 + coeffs.a1 * E + coeffs.a2 * N
                     + coeffs.a3 * E**2 + coeffs.a4 * E * N + coeffs.a5 * N**2)
            utm_n = (coeffs.b0 + coeffs.b1 * E + coeffs.b2 * N
                     + coeffs.b3 * E**2 + coeffs.b4 * E * N + coeffs.b5 * N**2)
        else:
            # Simple affine transformation
            utm_e = coeffs.a0 + coeffs.a1 * E + coeffs.a2 * N
            utm_n = coeffs.b0 + coeffs.b1 * E + coeffs.b2 * N

        return TransformResult(
            input_easting=easting,
            input_northing=northing,
            input_zone="polynomial",
            output_easting=round(utm_e, 3),
            output_northing=round(utm_n, 3),
            output_epsg=32737,  # Assume UTM 37S
            method="polynomial",
            accuracy_m=None,  # Depends on control point quality
        )

    def load_polynomial_from_json(self, json_data: Dict) -> PolynomialCoefficients:
        """Load polynomial coefficients from JSON (e.g., from API request)."""
        return PolynomialCoefficients(
            a0=json_data.get("a0", 0.0),
            a1=json_data.get("a1", 1.0),
            a2=json_data.get("a2", 0.0),
            a3=json_data.get("a3", 0.0),
            a4=json_data.get("a4", 0.0),
            a5=json_data.get("a5", 0.0),
            b0=json_data.get("b0", 0.0),
            b1=json_data.get("b1", 0.0),
            b2=json_data.get("b2", 1.0),
            b3=json_data.get("b3", 0.0),
            b4=json_data.get("b4", 0.0),
            b5=json_data.get("b5", 0.0),
        )

    def load_polynomial_from_excel(self, filepath: str) -> PolynomialCoefficients:
        """
        Extract polynomial coefficients from an Excel file.
        Looks for cells labeled with coefficient names (a0-a5, b0-b5).
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl required for Excel parsing")

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        coeffs = PolynomialCoefficients()
        coeff_map = {}

        for row in ws.iter_rows(values_only=True):
            for cell in row:
                if isinstance(cell, str):
                    key = cell.strip().lower()
                    if key in [f"a{i}" for i in range(6)] + [f"b{i}" for i in range(6)]:
                        # Find the value in adjacent cell
                        row_idx = ws[cell].row
                        col_idx = ws[cell].column
                        val_cell = ws.cell(row=row_idx, column=col_idx + 1)
                        if val_cell.value is not None:
                            coeff_map[key] = float(val_cell.value)

        # Also try a table format: Column A = label, Column B = value
        if not coeff_map:
            for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                if row[0] and row[1]:
                    label = str(row[0]).strip().lower()
                    try:
                        val = float(row[1])
                        coeff_map[label] = val
                    except (ValueError, TypeError):
                        pass

        if coeff_map:
            coeffs.a0 = coeff_map.get("a0", coeffs.a0)
            coeffs.a1 = coeff_map.get("a1", coeffs.a1)
            coeffs.a2 = coeff_map.get("a2", coeffs.a2)
            coeffs.a3 = coeff_map.get("a3", coeffs.a3)
            coeffs.a4 = coeff_map.get("a4", coeffs.a4)
            coeffs.a5 = coeff_map.get("a5", coeffs.a5)
            coeffs.b0 = coeff_map.get("b0", coeffs.b0)
            coeffs.b1 = coeff_map.get("b1", coeffs.b1)
            coeffs.b2 = coeff_map.get("b2", coeffs.b2)
            coeffs.b3 = coeff_map.get("b3", coeffs.b3)
            coeffs.b4 = coeff_map.get("b4", coeffs.b4)
            coeffs.b5 = coeff_map.get("b5", coeffs.b5)

        wb.close()
        return coeffs

    # ══════════════════════════════════════════════════════════════════
    #  APPROACH 3: Helmert 7-parameter transformation
    # ══════════════════════════════════════════════════════════════════

    def cassini_to_utm_helmert(
        self,
        easting: float,
        northing: float,
        params: Dict[str, float],
        source_epsg: int = 42009,  # Clarke 1880 RGS ( geographic)
        target_epsg: int = 4326,   # WGS84
    ) -> Optional[TransformResult]:
        """
        Convert via Helmert 7-parameter transformation.

        params should contain:
            dx, dy, dz: translation (metres)
            rx, ry, rz: rotation (arcseconds)
            s: scale (ppm)
        """
        if not self.pyproj_available:
            return None

        try:
            # Build towgs84 string
            towgs84 = (
                f"{params['dx']},{params['dy']},{params['dz']},"
                f"{params.get('rx', 0)},{params.get('ry', 0)},{params.get('rz', 0)},"
                f"{params.get('s', 0)}"
            )

            # Source CRS with Helmert shift
            source_crs = CRS.from_proj4(
                f"+proj=longlat +ellps=clrk80 +towgs84={towgs84} +no_defs"
            )
            target_crs = CRS.from_epsg(32737)  # UTM 37S

            # First: Cassini → Geographic
            # We need to know the geographic coords to do the datum shift
            # Build a Cassini CRS with the towgs84
            cassini_proj4 = (
                f"+proj=cass +lat_0=0 +lon_0=38 +k=1 "
                f"+x_0=700000 +y_0=5000000 "
                f"+ellps=clrk80 +towgs84={towgs84} +units=m +no_defs"
            )
            cassini_crs = CRS.from_proj4(cassini_proj4)

            # Transform directly to UTM
            transformer = Transformer.from_crs(cassini_crs, target_crs, always_xy=True)
            utm_e, utm_n = transformer.transform(easting, northing)

            # Get geographic coords
            geo_transformer = Transformer.from_crs(
                cassini_crs, "EPSG:4326", always_xy=True
            )
            lon, lat = geo_transformer.transform(easting, northing)

            return TransformResult(
                input_easting=easting,
                input_northing=northing,
                input_zone="helmert",
                output_easting=round(utm_e, 3),
                output_northing=round(utm_n, 3),
                output_epsg=32737,
                latitude=round(lat, 8),
                longitude=round(lon, 8),
                method="helmert_7param",
                accuracy_m=0.1,  # Sub-metre with good control points
            )
        except Exception:
            return None

    # ══════════════════════════════════════════════════════════════════
    #  Bulk transformation from Excel/CSV
    # ══════════════════════════════════════════════════════════════════

    def transform_excel(
        self,
        filepath: str,
        method: str = "geodetic",
        zone: str = "zone_3",
        utm_zone: int = 37,
        polynomial_coeffs: Optional[Dict] = None,
        input_columns: Optional[Dict[str, str]] = None,
    ) -> Dict:
        """
        Transform coordinates from an Excel file.

        Args:
            filepath: Path to Excel file
            method: 'geodetic', 'polynomial', or 'helmert'
            zone: Cassini zone (for geodetic method)
            utm_zone: Target UTM zone
            polynomial_coeffs: Coefficients for polynomial method
            input_columns: Column name mapping, e.g.
                {"easting": "Cassini East", "northing": "Cassini North"}

        Returns:
            Dict with results array and statistics
        """
        if not OPENPYXL_AVAILABLE:
            return {"error": "openpyxl not installed. Run: pip install openpyxl"}

        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active

        # Auto-detect columns
        headers = [str(cell.value).strip().lower() if cell.value else ""
                   for cell in ws[1]]

        col_map = input_columns or {}
        east_col = self._find_column(headers, col_map.get("easting", []),
                                     ["easting", "east", "x", "e_cassini", "c_east"])
        north_col = self._find_column(headers, col_map.get("northing", []),
                                      ["northing", "north", "y", "n_cassini", "c_north"])

        if east_col is None or north_col is None:
            wb.close()
            return {
                "error": "Could not find easting/northing columns. "
                         "Headers found: " + str(headers),
                "hint": "Use input_columns parameter to specify column names",
            }

        # Load polynomial coefficients from a second sheet if provided
        coeffs = None
        if method == "polynomial":
            if polynomial_coeffs:
                coeffs = self.load_polynomial_from_json(polynomial_coeffs)
            elif len(wb.sheetnames) > 1:
                coeffs = self.load_polynomial_from_excel(filepath)

        results = []
        errors = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                E = float(row[east_col])
                N = float(row[north_col])

                if method == "geodetic":
                    result = self.cassini_to_utm_geodetic(E, N, zone, utm_zone)
                elif method == "polynomial":
                    if not coeffs:
                        errors.append({"row": row_idx, "error": "No polynomial coefficients"})
                        continue
                    result = self.cassini_to_utm_polynomial(E, N, coeffs)
                elif method == "helmert":
                    result = self.cassini_to_utm_helmert(E, N, KENYA_DATUM_SHIFTS["arc_1960"])
                else:
                    result = self.cassini_to_utm_geodetic(E, N, zone, utm_zone)

                if result:
                    results.append({
                        "row": row_idx,
                        "input_easting": result.input_easting,
                        "input_northing": result.input_northing,
                        "output_easting": result.output_easting,
                        "output_northing": result.output_northing,
                        "latitude": result.latitude,
                        "longitude": result.longitude,
                    })
                else:
                    errors.append({"row": row_idx, "error": "Transformation failed"})

            except (ValueError, TypeError) as e:
                errors.append({"row": row_idx, "error": str(e)})

        wb.close()

        return {
            "total_rows": len(results) + len(errors),
            "successful": len(results),
            "failed": len(errors),
            "results": results,
            "errors": errors,
            "method": method,
            "zone": zone,
            "utm_zone": utm_zone,
        }

    def _find_column(self, headers: List[str], explicit_names: List[str],
                     fallback_names: List[str]) -> Optional[int]:
        """Find column index by name matching."""
        # Try explicit names first
        for name in explicit_names:
            for idx, h in enumerate(headers):
                if name.lower() in h:
                    return idx

        # Try fallback names
        for name in fallback_names:
            for idx, h in enumerate(headers):
                if name in h:
                    return idx

        return None

    # ══════════════════════════════════════════════════════════════════
    #  Utility: Auto-detect UTM zone from coordinates
    # ══════════════════════════════════════════════════════════════════

    @staticmethod
    def detect_utm_zone(longitude: float) -> int:
        """Determine UTM zone from longitude."""
        return int((longitude + 180) / 6) + 1

    @staticmethod
    def detect_cassini_zone(longitude: float) -> str:
        """Best-guess Cassini zone from longitude."""
        if longitude < 35.0:
            return "zone_1"
        elif longitude < 37.0:
            return "zone_2"
        elif longitude < 39.0:
            return "zone_3"
        else:
            return "zone_4"

    def get_available_zones(self) -> Dict:
        """Return all supported zones and their parameters."""
        return {
            "cassini_zones": KENYA_CASSINI_ZONES,
            "datum_shifts": {k: v["name"] for k, v in KENYA_DATUM_SHIFTS.items()},
            "methods": ["geodetic", "polynomial", "helmert"],
        }
